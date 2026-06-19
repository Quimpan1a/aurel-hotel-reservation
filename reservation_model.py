import sqlite3
import csv
import hashlib
import os
from typing import Optional
from dataclasses import dataclass


DB_PATH = "hotel_reservations.db"

@dataclass
class User:
    id: Optional[int] = None
    username: str = ""
    password_hash: str = ""
    role: str = "staff"         
    full_name: str = ""
    is_active: bool = True

    @staticmethod
    def hash_password(password: str) -> str:
        """SHA-256 with a fixed salt prefix (simple, no bcrypt dependency)."""
        salted = f"aurel_hotel_salt_{password}"
        return hashlib.sha256(salted.encode()).hexdigest()

    @staticmethod
    def from_row(row) -> "User":
        return User(
            id=row[0],
            username=row[1],
            password_hash=row[2],
            role=row[3],
            full_name=row[4] if len(row) > 4 else "",
            is_active=bool(row[5]) if len(row) > 5 else True,
        )

@dataclass
class Reservation:
    id: Optional[int] = None
    guest_name: str = ""
    room_number: str = ""
    room_type: str = ""
    check_in: str = ""
    check_out: str = ""
    guests: int = 1
    total_price: float = 0.0
    status: str = "Confirmed"
    notes: str = ""

    def to_tuple(self):
        return (
            self.guest_name,
            self.room_number,
            self.room_type,
            self.check_in,
            self.check_out,
            self.guests,
            self.total_price,
            self.status,
            self.notes,
        )

    @staticmethod
    def from_row(row):
        return Reservation(
            id=row[0],
            guest_name=row[1],
            room_number=row[2],
            room_type=row[3],
            check_in=row[4],
            check_out=row[5],
            guests=int(row[6]),
            total_price=float(row[7]),
            status=row[8],
            notes=row[9] if len(row) > 9 else "",
        )


class ReservationDB:

    EXPECTED_COLUMNS = {
        "id", "guest_name", "room_number", "room_type",
        "check_in", "check_out", "guests", "total_price", "status", "notes",
    }

    def __init__(self):
        self._init_db()

    def _conn(self):
        return sqlite3.connect(DB_PATH)

    def _init_db(self):
        with self._conn() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS reservations (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    guest_name  TEXT    NOT NULL DEFAULT '',
                    room_number TEXT    NOT NULL DEFAULT '',
                    room_type   TEXT    NOT NULL DEFAULT '',
                    check_in    TEXT    NOT NULL DEFAULT '',
                    check_out   TEXT    NOT NULL DEFAULT '',
                    guests      INTEGER NOT NULL DEFAULT 1,
                    total_price REAL    NOT NULL DEFAULT 0.0,
                    status      TEXT    NOT NULL DEFAULT 'Confirmed',
                    notes       TEXT             DEFAULT ''
                )
            """)

           
            c.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    username      TEXT    NOT NULL UNIQUE,
                    password_hash TEXT    NOT NULL,
                    role          TEXT    NOT NULL DEFAULT 'staff',
                    full_name     TEXT             DEFAULT '',
                    is_active     INTEGER NOT NULL DEFAULT 1
                )
            """)
            c.commit()

        
            user_cols = {
                row[1]
                for row in c.execute("PRAGMA table_info(users)").fetchall()
            }
            if "full_name" not in user_cols:
                c.execute("ALTER TABLE users ADD COLUMN full_name TEXT DEFAULT ''")
            if "is_active" not in user_cols:
                c.execute("ALTER TABLE users ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
            c.commit()

           
            for uname, pw, role, fname in [
                ("admin", "admin123", "admin", "System Administrator"),
                ("staff", "staff123", "staff", "Front Desk Staff"),
            ]:
                pw_hash = User.hash_password(pw)
                exists = c.execute(
                    "SELECT id FROM users WHERE username=?", (uname,)
                ).fetchone()
                if exists:
                    c.execute(
                        "UPDATE users SET password_hash=?, role=?, full_name=?, is_active=1 "
                        "WHERE username=?",
                        (pw_hash, role, fname, uname),
                    )
                else:
                    c.execute(
                        "INSERT INTO users (username, password_hash, role, full_name, is_active) "
                        "VALUES (?,?,?,?,1)",
                        (uname, pw_hash, role, fname),
                    )
            c.commit()

            existing_cols = {
                row[1]
                for row in c.execute("PRAGMA table_info(reservations)").fetchall()
            }
            missing = self.EXPECTED_COLUMNS - existing_cols
            if missing:
                c.execute("ALTER TABLE reservations RENAME TO _reservations_backup")
                c.execute("""
                    CREATE TABLE reservations (
                        id          INTEGER PRIMARY KEY AUTOINCREMENT,
                        guest_name  TEXT    NOT NULL DEFAULT '',
                        room_number TEXT    NOT NULL DEFAULT '',
                        room_type   TEXT    NOT NULL DEFAULT '',
                        check_in    TEXT    NOT NULL DEFAULT '',
                        check_out   TEXT    NOT NULL DEFAULT '',
                        guests      INTEGER NOT NULL DEFAULT 1,
                        total_price REAL    NOT NULL DEFAULT 0.0,
                        status      TEXT    NOT NULL DEFAULT 'Confirmed',
                        notes       TEXT             DEFAULT ''
                    )
                """)
                shared = sorted(existing_cols & self.EXPECTED_COLUMNS)
                if shared:
                    cols = ", ".join(shared)
                    c.execute(
                        f"INSERT INTO reservations ({cols}) "
                        f"SELECT {cols} FROM _reservations_backup"
                    )
                c.execute("DROP TABLE _reservations_backup")
                c.commit()


    def authenticate(self, username: str, password: str) -> Optional[User]:
        """Returns User on success, None on failure."""
        pw_hash = User.hash_password(password)
        with self._conn() as c:
            row = c.execute(
                "SELECT * FROM users WHERE username=? AND password_hash=? AND is_active=1",
                (username, pw_hash),
            ).fetchone()
        return User.from_row(row) if row else None

    def get_all_users(self) -> list:
        with self._conn() as c:
            rows = c.execute("SELECT * FROM users ORDER BY id").fetchall()
        return [User.from_row(r) for r in rows]

    def get_user_by_id(self, uid: int) -> Optional[User]:
        with self._conn() as c:
            row = c.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        return User.from_row(row) if row else None

    def add_user(self, user: User) -> int:
        with self._conn() as c:
            cur = c.execute(
                "INSERT INTO users (username, password_hash, role, full_name, is_active) "
                "VALUES (?,?,?,?,?)",
                (user.username, user.password_hash, user.role,
                 user.full_name, int(user.is_active)),
            )
            c.commit()
            return cur.lastrowid

    def update_user(self, user: User):
        with self._conn() as c:
            c.execute(
                "UPDATE users SET username=?, role=?, full_name=?, is_active=? WHERE id=?",
                (user.username, user.role, user.full_name, int(user.is_active), user.id),
            )
            c.commit()

    def update_user_password(self, uid: int, new_hash: str):
        with self._conn() as c:
            c.execute(
                "UPDATE users SET password_hash=? WHERE id=?",
                (new_hash, uid),
            )
            c.commit()

    def delete_user(self, uid: int):
        with self._conn() as c:
            c.execute("DELETE FROM users WHERE id=?", (uid,))
            c.commit()

    def username_exists(self, username: str, exclude_id: int = None) -> bool:
        with self._conn() as c:
            if exclude_id:
                row = c.execute(
                    "SELECT id FROM users WHERE username=? AND id!=?",
                    (username, exclude_id),
                ).fetchone()
            else:
                row = c.execute(
                    "SELECT id FROM users WHERE username=?", (username,)
                ).fetchone()
        return row is not None


    def add(self, r: Reservation) -> int:
        with self._conn() as c:
            cur = c.execute(
                """INSERT INTO reservations
                   (guest_name, room_number, room_type, check_in, check_out,
                    guests, total_price, status, notes)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                r.to_tuple(),
            )
            c.commit()
            return cur.lastrowid

    def update(self, r: Reservation):
        with self._conn() as c:
            c.execute(
                """UPDATE reservations
                   SET guest_name=?, room_number=?, room_type=?,
                       check_in=?, check_out=?, guests=?,
                       total_price=?, status=?, notes=?
                   WHERE id=?""",
                (*r.to_tuple(), r.id),
            )
            c.commit()

    def delete(self, rid: int):
        with self._conn() as c:
            c.execute("DELETE FROM reservations WHERE id=?", (rid,))
            c.commit()

    def get_all(self) -> list:
        with self._conn() as c:
            rows = c.execute(
                "SELECT * FROM reservations ORDER BY id DESC"
            ).fetchall()
        return [Reservation.from_row(r) for r in rows]

    def get_by_id(self, rid: int) -> Optional[Reservation]:
        with self._conn() as c:
            row = c.execute(
                "SELECT * FROM reservations WHERE id=?", (rid,)
            ).fetchone()
        return Reservation.from_row(row) if row else None

    def search(self, query: str) -> list:
        q = f"%{query}%"
        with self._conn() as c:
            rows = c.execute(
                """SELECT * FROM reservations
                   WHERE guest_name  LIKE ?
                      OR room_number LIKE ?
                      OR room_type   LIKE ?
                      OR status      LIKE ?
                   ORDER BY id DESC""",
                (q, q, q, q),
            ).fetchall()
        return [Reservation.from_row(r) for r in rows]


    def binary_search_by_name(self, name: str) -> list:
        all_records = sorted(self.get_all(), key=lambda r: r.guest_name.lower())
        if not all_records:
            return []
        target = name.strip().lower()
        lo, hi = 0, len(all_records) - 1
        found = -1
        while lo <= hi:
            mid = (lo + hi) // 2
            mid_val = all_records[mid].guest_name.lower()
            if mid_val == target:
                found = mid
                break
            elif mid_val < target:
                lo = mid + 1
            else:
                hi = mid - 1
        if found == -1:
            return []
        results = [all_records[found]]
        i = found - 1
        while i >= 0 and all_records[i].guest_name.lower() == target:
            results.append(all_records[i])
            i -= 1
        i = found + 1
        while i < len(all_records) and all_records[i].guest_name.lower() == target:
            results.append(all_records[i])
            i += 1
        return results


    def import_csv(self, filepath: str):
        REQUIRED = {"guest_name", "room_number", "room_type", "check_in", "check_out"}
        imported = 0
        errors = []
        try:
            with open(filepath, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return 0, ["CSV file has no headers."]
                headers = {h.strip().lower() for h in reader.fieldnames}
                missing = REQUIRED - headers
                if missing:
                    return 0, [f"Missing columns: {', '.join(sorted(missing))}"]
                for i, row in enumerate(reader, start=2):
                    try:
                        r = Reservation(
                            guest_name=row.get("guest_name", "").strip(),
                            room_number=row.get("room_number", "").strip(),
                            room_type=row.get("room_type", "").strip(),
                            check_in=row.get("check_in", "").strip(),
                            check_out=row.get("check_out", "").strip(),
                            guests=int(row.get("guests", 1) or 1),
                            total_price=float(row.get("total_price", 0) or 0),
                            status=(row.get("status", "").strip() or "Confirmed"),
                            notes=row.get("notes", "").strip(),
                        )
                        if not r.guest_name or not r.room_number:
                            errors.append(f"Row {i}: guest_name or room_number is empty.")
                            continue
                        self.add(r)
                        imported += 1
                    except Exception as exc:
                        errors.append(f"Row {i}: {exc}")
        except FileNotFoundError:
            return 0, ["File not found."]
        except Exception as exc:
            return 0, [str(exc)]
        return imported, errors


    def export_excel(self, filepath: str, reservations=None):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Run:  pip install openpyxl")

        data = reservations if reservations is not None else self.get_all()
        STATUS_CLR = {
            "Confirmed":   "27AE60",
            "Pending":     "E67E22",
            "Checked In":  "2980B9",
            "Checked Out": "7F8C8D",
            "Cancelled":   "C0392B",
        }
        thin   = Side(style="thin", color="BDC3C7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservations"

        ws.merge_cells("A1:J1")
        tc = ws["A1"]
        tc.value = "Hotel Reservation Report"
        tc.font  = Font(bold=True, size=15, color="1A2744")
        tc.alignment = center
        tc.fill = PatternFill("solid", fgColor="D6E4FF")
        ws.row_dimensions[1].height = 28

        HEADERS = [
            ("ID", 6), ("Guest Name", 22), ("Room No.", 10), ("Room Type", 14),
            ("Check-In", 12), ("Check-Out", 12), ("Guests", 8),
            ("Total Price ($)", 16), ("Status", 14), ("Notes", 28),
        ]
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="1A2744")
        for col, (label, width) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 22

        alt_fill = PatternFill("solid", fgColor="EEF3FF")
        for ri, res in enumerate(data, start=3):
            row_vals = [
                res.id, res.guest_name, res.room_number, res.room_type,
                res.check_in, res.check_out, res.guests,
                round(res.total_price, 2), res.status, res.notes,
            ]
            use_alt = (ri % 2 == 0)
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.border = border
                cell.alignment = left if col in (2, 10) else center
                if col == 9:
                    clr = STATUS_CLR.get(str(val), "888888")
                    cell.fill = PatternFill("solid", fgColor=clr)
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                elif use_alt:
                    cell.fill = alt_fill
            ws.row_dimensions[ri].height = 18

        last = len(data) + 3
        tot_fill = PatternFill("solid", fgColor="D6E4FF")
        tot_font = Font(bold=True, color="1A2744", size=11)
        for col in range(1, 11):
            cell = ws.cell(row=last, column=col)
            cell.fill = tot_fill
            cell.border = border
            cell.font = tot_font
            cell.alignment = center
        ws.cell(row=last, column=1, value="TOTAL")
        ws.cell(row=last, column=7, value=sum(r.guests for r in data))
        ws.cell(row=last, column=8, value=round(sum(r.total_price for r in data), 2))
        ws.row_dimensions[last].height = 20
        ws.freeze_panes = "A3"
        wb.save(filepath)